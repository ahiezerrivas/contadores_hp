"""
Importa los reportes mensuales de contadores (Excel) hacia MonthlyCounterEntry.

Los archivos esperados son del tipo:
    Contador de Pagina Torre Julio del 01 al 31.xlsx
    Contador de Pagina Oficina Julio del 01 al 31.xlsx

Con columnas (en cualquier orden, se detectan por nombre de encabezado):
    Region, Nombre Oficina, Piso, Nombre de Host Sede, Asignada o Ubicacion,
    DisplayName, IPv4Address, SerialNumber, Contador del Mes Anterior,
    Contador Semana 1, Contador Final Semana 1, Contador Semana 2,
    Contador Final Semana 2, Contador Semana 3, Contador Final Semana 3,
    Contador Semana 4, Contador Final Semana 4, Contador Semana 5,
    Contador Final Semana 5, Contador Mensual, Equipos con contadores en 0,
    Fecha, Observaciones, Status Impresora

Uso:
    python manage.py import_counters "Contador de Pagina Torre Julio del 01 al 31.xlsx"
    python manage.py import_counters archivo.xlsx --region Torre --period jul-26
    python manage.py import_counters archivo.xlsx --sheet "Hoja1" --dry-run
"""

import datetime
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q
from openpyxl import load_workbook

from devices.models import Impresora, MonthlyCounterEntry, Oficina
from devices.utils import SPANISH_MONTHS

# Campos que ahora viven en el modelo Impresora (antes estaban en MonthlyCounterEntry).
# Mapea el nombre de campo usado en este importador -> nombre real del campo en Impresora.
IMPRESORA_FIELD_MAP = {
    "display_name": "name",
    "ip_address": "ip_address",
    "serial_number": "serial_number",
    "printer_status": "status",
}

# Encabezado normalizado -> nombre de campo del modelo.
HEADER_TO_FIELD = {
    "region": "region",
    "nombre oficina": "office_name",
    "piso": "floor",
    "asignada o ubicacion": "location",
    "displayname": "display_name",
    "display name": "display_name",
    "ipv4address": "ip_address",
    "ipv4 address": "ip_address",
    "serialnumber": "serial_number",
    "serial number": "serial_number",
    "serial": "serial_number",
    "contador del mes anterior": "previous_month_counter",
    "contador mes anterior": "previous_month_counter",
    "contador semana 1": "week1_counter",
    "contador final semana 1": "week1_final",
    "contador semana 2": "week2_counter",
    "contador final semana 2": "week2_final",
    "contador semana 3": "week3_counter",
    "contador final semana 3": "week3_final",
    "contador semana 4": "week4_counter",
    "contador final semana 4": "week4_final",
    "contador semana 5": "week5_counter",
    "contador final semana 5": "week5_final",
    "contador mensual": "monthly_counter",
    "equipos con contadores en 0": "zero_counter_devices",
    "fecha": "period",
    "observaciones": "observations",
    "status impresora": "printer_status",
}

NUMERIC_FIELDS = {
    "previous_month_counter",
    "week1_counter", "week1_final",
    "week2_counter", "week2_final",
    "week3_counter", "week3_final",
    "week4_counter", "week4_final",
    "week5_counter", "week5_final",
    "monthly_counter",
    "zero_counter_devices",
}

TEXT_FIELDS = {
    "region", "floor", "location",
    "display_name", "ip_address", "serial_number", "observations",
    "printer_status",
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def to_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip()
    if not text or text in {"-", "N/A", "n/a"}:
        return None
    digits = re.sub(r"[^\d-]", "", text)
    if not digits or digits == "-":
        return None
    return int(digits)


def format_period(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return f"{SPANISH_MONTHS[value.month]}-{str(value.year)[2:]}"
    return str(value).strip()


class Command(BaseCommand):
    help = "Importa un reporte mensual de contadores (Excel) a MonthlyCounterEntry."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Ruta al archivo .xlsx")
        parser.add_argument("--sheet", default=None, help="Nombre de la hoja (por defecto la activa).")
        parser.add_argument(
            "--region",
            default=None,
            help="Forzar el valor de Region para todas las filas (ej. Torre, Oficina).",
        )
        parser.add_argument(
            "--period",
            default=None,
            help="Forzar el periodo (ej. jul-26) para todas las filas, en vez de usar la columna Fecha.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Muestra cuantas filas se importarian sin guardar nada en la base de datos.",
        )

    def handle(self, *args, **options):
        path = options["path"]
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontro el archivo: {path}")
        except Exception as exc:
            raise CommandError(f"No se pudo abrir el archivo: {exc}")

        sheet_name = options.get("sheet")
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise CommandError("El archivo esta vacio.")

        field_by_column = {}
        for idx, header in enumerate(header_row):
            normalized = normalize_header(header)
            field = HEADER_TO_FIELD.get(normalized)
            if field:
                field_by_column[idx] = field

        if not field_by_column:
            raise CommandError(
                "No se reconocio ninguna columna del encabezado. "
                "Verifica que la primera fila tenga los nombres esperados."
            )

        override_region = options.get("region")
        override_period = options.get("period")
        dry_run = options.get("dry_run")

        created, updated, skipped = 0, 0, 0

        for row in rows_iter:
            if row is None or all(cell is None for cell in row):
                continue

            data = {}
            for idx, field in field_by_column.items():
                value = row[idx] if idx < len(row) else None
                if field in NUMERIC_FIELDS:
                    data[field] = to_int(value)
                elif field == "period":
                    data[field] = format_period(value)
                else:
                    data[field] = ("" if value is None else str(value).strip())

            if override_region:
                data["region"] = override_region
            if override_period:
                data["period"] = override_period

            ip_address = data.get("ip_address", "")
            display_name = data.get("display_name", "")
            serial_number = data.get("serial_number", "")

            if not (ip_address or display_name or serial_number):
                skipped += 1
                continue

            office_name_text = data.pop("office_name", "")
            if office_name_text:
                office, _ = Oficina.objects.get_or_create(
                    name=office_name_text, defaults={"status": "Activo"}
                )
                data["office"] = office
            else:
                data["office"] = None

            data["source_file"] = path

            impresora_data = {
                IMPRESORA_FIELD_MAP[field]: data.pop(field)
                for field in list(IMPRESORA_FIELD_MAP)
                if field in data
            }

            if dry_run:
                created += 1
                continue

            impresora = None
            imp_q = None
            if serial_number:
                imp_q = Q(serial_number=serial_number)
            if ip_address:
                imp_q = imp_q | Q(ip_address=ip_address) if imp_q else Q(ip_address=ip_address)
            if imp_q:
                impresora = Impresora.objects.filter(imp_q).first()

            if impresora:
                for attr, value in impresora_data.items():
                    if value:
                        setattr(impresora, attr, value)
                impresora.save()
            elif any(impresora_data.values()):
                impresora = Impresora.objects.create(**impresora_data)

            data["impresora"] = impresora

            if impresora:
                existing = MonthlyCounterEntry.objects.filter(
                    impresora=impresora, period=data.get("period", "")
                ).first()
                if existing:
                    for attr, value in data.items():
                        setattr(existing, attr, value)
                    existing.save()
                    updated += 1
                else:
                    MonthlyCounterEntry.objects.create(**data)
                    created += 1
            else:
                MonthlyCounterEntry.objects.create(**data)
                created += 1

        workbook.close()

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    f"[dry-run] {created} filas se importarian, {skipped} filas omitidas (vacias)."
                )
            )
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f"Import completado: {created} creados, {updated} actualizados, "
                    f"{skipped} omitidos (vacios)."
                )
            )
